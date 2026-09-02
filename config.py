import os
from openpyxl.styles import PatternFill

# ============================================================
# EXCEL FILE SETTINGS
# ============================================================

# Input workbook
INPUT_FILE = "MTN_DATA.xlsx"

# Worksheet names
ITU_SHEET = "for ITU"
SERVICE_SHEET = "Service Fixe PP Public"

# ============================================================
# OUTPUT SETTINGS
# ============================================================

# Folder where generated ITU notices will be saved
OUTPUT_FOLDER = "ITU_NOTICES"
os.makedirs(OUTPUT_FOLDER, exist_ok=True) # Create the folder if it doesn't already exist

# ============================================================
# EXCEL ROW COLOURS
# ============================================================

# Successfully processed row
GREEN_FILL = PatternFill(
    start_color="00FF00",
    end_color="00FF00",
    fill_type="solid"
)

# Duplicate notice
BLUE_FILL = PatternFill(
    start_color="0000FF",
    end_color="0000FF",
    fill_type="solid"
)

# Error encountered while processing
RED_FILL = PatternFill(
    start_color="FF0000",
    end_color="FF0000",
    fill_type="solid"
)

# ============================================================
# NOTICE CONSTANTS
# ============================================================

ADMINISTRATION = "GHA"
NOTICE_TYPE = "T11"
FRAGMENT = "NTFD_RR"
ACTION = "ADD"
EMAIL = "Kweku.inkoom@nca.org.gh"
CHARSET = "ISO-8859-1"
COUNTRY = "GHA"
ADDRESS_CODE = "B"
STATION_CLASS = "FX"
NATIONAL_SERVICE = "OT"
PROVISION = "RR11.2"
GAIN_TYPE = "I"
POWER_TYPE = "I"
POWER_AXIS = "X"
ANTENNA_DIRECTION = "D"
EMISSION_CLASS = "W9WWW"